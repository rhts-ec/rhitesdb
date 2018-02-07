from django.shortcuts import render, get_object_or_404, render_to_response, redirect
from django.db.models import Avg, Case, Count, F, Max, Min, Prefetch, Q, Sum, When
from django.db.models import Value, CharField
from django.db.models.functions import Substr
from django.contrib.auth.decorators import login_required
from django.http import Http404
from django.template import RequestContext
from django.core.urlresolvers import reverse

from datetime import date
from decimal import Decimal
from itertools import groupby, tee, chain, product

from . import dateutil, grabbag
from .grabbag import default_zero, all_not_none

from .models import DataElement, OrgUnit, DataValue, ValidationRule, SourceDocument
from .forms import SourceDocumentForm, DataElementAliasForm

@login_required
def index(request):
    context = {
        'validation_rules': ValidationRule.objects.all().values_list('id', 'name')
    }
    return render(request, 'cannula/index.html', context)

@login_required
def data_elements(request):
    data_elements = DataElement.objects.order_by('name').all()
    return render(request, 'cannula/data_element_listing.html', {'data_elements': data_elements})

# avoid strange behaviour from itertools.groupby by evaluating all the group iterables as lists
def groupbylist(*args, **kwargs):
    return [[k, list(g)] for k, g in groupby(*args, **kwargs)]

def filter_empty_rows(grouped_vals):
    for row in grouped_vals:
        row_heading, row_values = row
        if any(v['numeric_sum'] for v in row_values):
            yield row

def month2quarter(month_num):
    return ((month_num-1)//3+1)

@login_required
def ipt_quarterly(request, output_format='HTML'):
    ipt_de_names = (
        '105-2.1 A6:First dose IPT (IPT1)',
        '105-2.1 A7:Second dose IPT (IPT2)',
    )

    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    # get IPT1 and IPT2 without subcategory disaggregation
    qs = DataValue.objects.what(*ipt_de_names).filter(quarter=filter_period)
    # use clearer aliases for the unwieldy names
    qs = qs.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'))
    qs = qs.annotate(period=F('quarter')) # TODO: review if this can still work with different periods
    qs = qs.order_by('district', 'subcounty', 'de_name', 'period')
    val_dicts = qs.values('district', 'subcounty', 'de_name', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    
    # all subcounties (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=2).annotate(district=F('parent__name'), subcounty=F('name'))
    ou_list = qs_ou.values_list('district', 'subcounty')

    def val_fun(row, col):
        return { 'district': row[0], 'subcounty': row[1], 'de_name': col, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, ipt_de_names, val_dicts, lambda x: (x['district'], x['subcounty']), lambda x: x['de_name'], val_fun)
    val_dicts = list(gen_raster)

    # get list of subcategories for IPT2
    qs_ipt_subcat = DataValue.objects.what('105-2.1 A7:Second dose IPT (IPT2)').order_by('category_combo__name').values_list('de_name', 'category_combo__name').distinct()
    subcategory_names = tuple(qs_ipt_subcat)

    # get IPT2 with subcategory disaggregation
    qs2 = DataValue.objects.what('105-2.1 A7:Second dose IPT (IPT2)').filter(quarter=filter_period)
    # use clearer aliases for the unwieldy names
    qs2 = qs2.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'))
    qs2 = qs2.annotate(period=F('quarter')) # TODO: review if this can still work with different periods
    qs2 = qs2.annotate(cat_combo=F('category_combo__name'))
    qs2 = qs2.order_by('district', 'subcounty', 'de_name', 'period', 'cat_combo')
    val_dicts2 = qs2.values('district', 'subcounty', 'de_name', 'period', 'cat_combo').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    def val_with_subcat_fun(row, col):
        district, subcounty = row
        de_name, subcategory = col
        return { 'district': district, 'subcounty': subcounty, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, subcategory_names, val_dicts2, lambda x: (x['district'], x['subcounty']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_dicts2 = list(gen_raster)

    # get expected pregnancies
    qs3 = DataValue.objects.what('Expected Pregnancies')
    # use clearer aliases for the unwieldy names
    qs3 = qs3.annotate(district=F('org_unit__parent__name'), subcounty=F('org_unit__name'))
    qs3 = qs3.annotate(period=F('year')) # TODO: review if this can still work with different periods
    qs3 = qs3.order_by('district', 'subcounty', 'de_name', 'period')
    val_dicts3 = qs3.values('district', 'subcounty', 'de_name', 'period').annotate(numeric_sum=(Sum('numeric_value')/4))

    gen_raster = grabbag.rasterize(ou_list, ('Expected Pregnancies',), val_dicts3, lambda x: (x['district'], x['subcounty']), lambda x: x['de_name'], val_fun)
    val_dicts3 = list(gen_raster)

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(chain(val_dicts3, val_dicts, val_dicts2), key=lambda x: (x['district'], x['subcounty'])), key=lambda x: (x['district'], x['subcounty']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))
    
    # calculate the IPT rate for the IPT1/IPT2 values (without subcategories)
    for _group in grouped_vals:
        (district_subcounty, (preg_val, *other_vals)) = _group
        if preg_val['de_name'] == 'Expected Pregnancies':
            for val in other_vals:
                if val['de_name'] in ipt_de_names and 'cat_combo' not in val:
                    pregnancies_per_annum = preg_val['numeric_sum']
                    if pregnancies_per_annum and pregnancies_per_annum != 0 and val['numeric_sum']:
                        val['ipt_rate'] = val['numeric_sum']*100/pregnancies_per_annum
                    else:
                        val['ipt_rate'] = None

    data_element_names = list()
    data_element_names.insert(0, ('Expected Pregnancies', None))
    for de_n in ipt_de_names:
        data_element_names.append((de_n, None))
        data_element_names.append(('%', None))
    data_element_names.extend(subcategory_names)

    if output_format == 'EXCEL':
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Color, PatternFill, Font, Border
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule

        wb = openpyxl.workbook.Workbook()
        ws = wb.create_sheet()
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        headers = ['District', 'Subcounty'] + data_element_names
        for i, name in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            if not isinstance(name, tuple):
                c.value = str(name)
            else:
                de, cat_combo = name
                if cat_combo is None:
                    c.value = str(de)
                else:
                    c.value = str(de) + '\n' + str(cat_combo)
        for i, g in enumerate(grouped_vals, start=2):
            (district, subcounty), g_val_list = g
            ws.cell(row=i, column=1, value=district)
            ws.cell(row=i, column=2, value=subcounty)
            offset = 0
            for j, g_val in enumerate(g_val_list, start=3):
                ws.cell(row=i, column=j+offset, value=g_val['numeric_sum'])
                if 'ipt_rate' in g_val:
                    offset += 1
                    ws.cell(row=i, column=j+offset, value=g_val['ipt_rate'])


        #ipt1_percent_range = 'E:E' # entire-column-range syntax doesn't work for conditional formatting
        # use old-school column/row limit as stand-in for entire row
        ipt1_percent_range = 'E1:E16384'
        ipt2_percent_range = 'G1:G16384'
        yellow_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
        green_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        rule_lt_71 = CellIsRule(operator='between', formula=['0','70'], stopIfTrue=True, fill=yellow_fill)
        rule_lt_71_unbounded = CellIsRule(operator='lessThan', formula=['71'], stopIfTrue=True, fill=yellow_fill)
        rule_ge_71 = CellIsRule(operator='between', formula=['71','100'], stopIfTrue=True, fill=green_fill)
        rule_ge_71_unbounded = CellIsRule(operator='greaterThanOrEqual', formula=['71'], stopIfTrue=True, fill=green_fill)
        rule_ignore_blanks = Rule(type="containsBlanks", stopIfTrue=True)
        ws.conditional_formatting.add(ipt1_percent_range, rule_ignore_blanks)
        ws.conditional_formatting.add(ipt1_percent_range, rule_lt_71)
        ws.conditional_formatting.add(ipt1_percent_range, rule_ge_71_unbounded)
        ws.conditional_formatting.add(ipt2_percent_range, rule_ignore_blanks)
        ws.conditional_formatting.add(ipt2_percent_range, rule_lt_71)
        ws.conditional_formatting.add(ipt2_percent_range, rule_ge_71_unbounded)


        response = HttpResponse(openpyxl.writer.excel.save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="malaria_ipt_scorecard.xlsx"'

        return response

    context = {
        'grouped_data': grouped_vals,
        'data_element_names': data_element_names,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
        'level_list': level_list,
    }

    if output_format == 'JSON':
        from django.http import JsonResponse
        
        return JsonResponse(context)

    return render(request, 'cannula/ipt_quarterly.html', context)

@login_required
def malaria_compliance(request):
    cases_de_names = (
        '105-1.3 OPD Malaria (Total)',
        '105-1.3 OPD Malaria Confirmed (Microscopic & RDT)',
    )

    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]

    if 'start_period' in request.GET and request.GET['start_period'] in PREV_5YR_QTRS and 'end_period' in request.GET and request.GET['end_period']:
        start_quarter = request.GET['start_period']
        end_quarter = request.GET['end_period']
    else: # default to "immediate preceding quarter" and "this quarter"
        if this_day.month <= 3:
            start_year = this_year - 1
            start_month = (this_day.month - 3 + 12)
            end_month = this_day.month
        else:
            start_year = this_year
            start_month = this_day.month - 3
            end_month = this_day.month
        start_quarter = '%d-Q%d' % (start_year, month2quarter(start_month))
        end_quarter = '%d-Q%d' % (this_year, month2quarter(end_month))

    periods = dateutil.get_quarters(start_quarter, end_quarter)
    if start_quarter == end_quarter:
        periods = periods[:1]
    
    # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=3).annotate(district=F('parent__parent__name'), subcounty=F('parent__name'), facility=F('name'))
    ou_list = qs_ou.values_list('district', 'subcounty', 'facility')

    # get data values without subcategory disaggregation
    qs = DataValue.objects.what(*cases_de_names)
    qs = qs.filter(quarter__gte=start_quarter).filter(quarter__lte=end_quarter)
    # use clearer aliases for the unwieldy names
    qs = qs.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs = qs.annotate(period=F('quarter')) # TODO: review if this can still work with different periods
    qs = qs.order_by('district', 'subcounty', 'facility', 'de_name', 'period')
    val_dicts = qs.values('district', 'subcounty', 'facility', 'de_name', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    def val_with_period_fun(row, col):
        district, subcounty, facility = row
        de_name, period = col
        return { 'district': district, 'subcounty': subcounty, 'facility': facility, 'period': period, 'de_name': de_name, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, tuple(product(cases_de_names, periods)), val_dicts, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['period']), val_with_period_fun)
    val_dicts = gen_raster

    # combine the data and group by district and subcounty
    grouped_vals = groupbylist(sorted(val_dicts, key=lambda x: (x['district'], x['subcounty'], x['facility'])), key=lambda x: (x['district'], x['subcounty'], x['facility']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    for _group in grouped_vals:
        (district_subcounty_facility, other_vals) = _group
        malaria_totals = dict()
        for val in other_vals:
            if val['de_name'] == cases_de_names[0]:
                malaria_totals[val['period']] = val['numeric_sum']
            elif val['de_name'] == cases_de_names[1]:
                total_cases = malaria_totals.get(val['period'], 0)
                confirmed_cases = val['numeric_sum']
                if confirmed_cases and total_cases and total_cases != 0:
                    confirmed_rate = confirmed_cases * 100 / total_cases
                    val['rdt_rate'] = confirmed_rate
                else:
                    val['rdt_rate'] = None

    data_element_names = list()
    for de_n in cases_de_names:
        data_element_names.append((de_n, None))

    context = {
        'grouped_data': grouped_vals,
        'data_element_names': data_element_names,
        'start_period': start_quarter,
        'end_period': end_quarter,
        'periods': periods,
        'period_desc': dateutil.DateSpan.fromquarter(start_quarter).combine(dateutil.DateSpan.fromquarter(end_quarter)).format_long(),
        'period_list': PREV_5YR_QTRS,
    }

    return render(request, 'cannula/malaria_compliance.html', context)

@login_required
def data_workflow_new(request):
    if request.method == 'POST':
        form = SourceDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('data_workflow_listing')
    else:
        form = SourceDocumentForm()

    context = {
        'form': form,
    }

    return render_to_response('cannula/data_workflow_new.html', context, context_instance=RequestContext(request))

@login_required
def data_workflow_detail(request):
    from .models import load_excel_to_datavalues, load_excel_to_validations

    if 'wf_id' in request.GET:
        src_doc_id = int(request.GET['wf_id'])
        src_doc = get_object_or_404(SourceDocument, id=src_doc_id)

        if request.method == 'POST':
            if 'load_values' in request.POST:
                all_values = load_excel_to_datavalues(src_doc)
                for site_name, site_vals in all_values.items():
                    DataValue.objects.bulk_create(site_vals)
            elif 'load_validations' in request.POST:
                load_excel_to_validations(src_doc)

            #TODO: redirect with to detail page?

        qs_vals = DataValue.objects.filter(source_doc__id=src_doc_id).values('id')
        doc_elements = DataElement.objects.filter(data_values__id__in=qs_vals).distinct('id')
        doc_rules = ValidationRule.objects.filter(data_elements__data_values__id__in=qs_vals).distinct('id')
        num_values = qs_vals.count()
    else:
        raise Http404("Workflow does not exist or workflow id is missing/invalid")

    context = {
        'srcdoc': src_doc,
        'num_values': num_values,
        'data_elements': doc_elements,
        'validation_rules': doc_rules,
    }

    return render(request, 'cannula/data_workflow_detail.html', context)

@login_required
def data_workflow_listing(request):
    # TODO: filter based on user who uploaded file?
    docs = SourceDocument.objects.all().annotate(num_values=Count('data_values'))
    docs = docs.order_by('uploaded_at')

    context = {
        'workflows': docs,
    }
    return render(request, 'cannula/data_workflow_listing.html', context)

def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

@login_required
def validation_rule(request):
    from django.db import connection
    cursor = connection.cursor()
    vr_id = int(request.GET['id'])
    vr = ValidationRule.objects.get(id=vr_id)
    cursor.execute('SELECT * FROM %s' % (vr.view_name(),))
    columns = [col[0] for col in cursor.description]
    de_name_map = dict()
    for de_id, de_name in DataElement.objects.all().values_list('id', 'name'):
        de_name_map['de_%d' % (de_id,)] = de_name
        columns = [c.replace('de_%d' % (de_id,), de_name) for c in columns] #TODO: can we include the alias, if there is one?
    results = dictfetchall(cursor)
    for r in results:
        r['data_values'] = dict()
        for k,v in r.items():
            if k in de_name_map:
                de_name = de_name_map[k]
                r['data_values'][de_name] = v
    if 'exclude_true' in request.GET:
        results = filter(lambda x: not x['de_calc_1'], results)
    context = {
        'results': results,
        'columns': columns,
        'rule': vr,
    }

    return render(request, 'cannula/validation_rule.html', context)

@login_required
def data_element_alias(request):
    if 'de_id' in request.GET:
        de_id = int(request.GET['de_id'])
        de = get_object_or_404(DataElement, id=de_id)

        if request.method == 'POST':
            form = DataElementAliasForm(request.POST, instance=de)
            if form.is_valid():
                form.save()
                de_url = '%s?wf_id=%d' % (reverse('data_workflow_detail'), int(request.GET['wf_id']))
                return redirect(de_url, de)
        else:
            form = DataElementAliasForm(instance=de)

        context = {
            'form': form,
        }
    else:
        raise Http404("Data Element does not exist or data element id is missing/invalid")

    return render_to_response('cannula/data_element_edit_alias.html', context, context_instance=RequestContext(request))

@login_required
def hts_by_site(request):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    hts_de_names = (
        '105-4 Number of clients who have been linked to care',
        '105-4 Number of Individuals who received HIV test results',
        '105-4 Number of Individuals who tested HIV positive',
    )
    hts_short_names = (
        'Linked',
        'Tested',
        'HIV+',
    )
    subcategory_names = ['(<15, Female)', '(<15, Male)', '(15+, Female)', '(15+, Male)']
    de_positivity_meta = list(product(hts_de_names, subcategory_names))

    qs_positivity = DataValue.objects.what(*hts_de_names).filter(quarter=filter_period)

    cc_lt_15 = ['18 Mths-<5 Years', '5-<10 Years', '10-<15 Years']
    cc_ge_15 = ['15-<19 Years', '19-<49 Years', '>49 Years']
    #TODO: cc_lt_15_f = CategoryCombo.from_cat_names(['Female', '<15']) gives a CategoryCombo instance that makes the Case statement clearer/safer
    qs_positivity = qs_positivity.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    qs_positivity = qs_positivity.exclude(cat_combo__iexact=None)

    qs_positivity = qs_positivity.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_positivity = qs_positivity.annotate(period=F('quarter'))
    qs_positivity = qs_positivity.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_positivity = qs_positivity.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    
    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=3).annotate(district=F('parent__parent__name'), subcounty=F('parent__name'), facility=F('name'))
    ou_list = list(qs_ou.values_list('district', 'subcounty', 'facility'))

    def val_with_subcat_fun(row, col):
        district, subcounty, facility = row
        de_name, subcategory = col
        return { 'district': district, 'subcounty': subcounty, 'facility': facility, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, de_positivity_meta, val_positivity, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_positivity2 = list(gen_raster)

    pmtct_mother_de_names = (
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.2a Women tested for HIV in labour (1st time this Pregnancy)',
        '105-2.3a Breastfeeding mothers tested for HIV(1st test)',
    )
    de_pmtct_mother_meta = list(product(('Pregnant Women tested for HIV',), (None,)))

    qs_pmtct_mother = DataValue.objects.what(*pmtct_mother_de_names).filter(quarter=filter_period)
    qs_pmtct_mother = qs_pmtct_mother.annotate(de_name=Value('Pregnant Women tested for HIV', output_field=CharField()))
    qs_pmtct_mother = qs_pmtct_mother.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother = qs_pmtct_mother.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_pmtct_mother = qs_pmtct_mother.annotate(period=F('quarter'))
    qs_pmtct_mother = qs_pmtct_mother.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother = qs_pmtct_mother.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_meta, val_pmtct_mother, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother2 = list(gen_raster)

    pmtct_mother_pos_de_names = (
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.2a Women testing HIV+ in labour (1st time this Pregnancy)',
        '105-2.2b Women testing HIV+ in labour (Retest this Pregnancy)',
        '105-2.3a Breastfeeding mothers newly testing HIV+(1st test)',
        '105-2.3b Breastfeeding mothers newly testing HIV+(retest)',
    )
    de_pmtct_mother_pos_meta = list(product(('Pregnant Women testing HIV+',), (None,)))

    qs_pmtct_mother_pos = DataValue.objects.what(*pmtct_mother_pos_de_names).filter(quarter=filter_period)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(de_name=Value('Pregnant Women testing HIV+', output_field=CharField()))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(period=F('quarter'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother_pos = qs_pmtct_mother_pos.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_pos_meta, val_pmtct_mother_pos, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother_pos2 = list(gen_raster)

    pmtct_child_de_names = (
        '105-2.4a Exposed Infants Tested for HIV Below 18 Months(by 1st PCR) ',
        '105-2.4b 1st DNA PCR result returned(HIV+)',
        '105-2.4b 2nd DNA PCR result returned(HIV+)',
        '105-2.1a Male partners received HIV test results in eMTCT(Total)',
        '105-2.1b Male partners received HIV test results in eMTCT(HIV+)',
    )
    pmtct_child_short_names = (
        'PMTCT INFANT HIV+',
        'PMTCT CHILD PCR1 HIV+',
        'PMTCT CHILD PCR2 HIV+',
        'PMTCT MALE PARTNERS TESTED',
        'PMTCT MALE PARTNERS HIV+',
    )
    de_pmtct_child_meta = list(product(pmtct_child_de_names, (None,)))

    qs_pmtct_child = DataValue.objects.what(*pmtct_child_de_names).filter(quarter=filter_period)
    qs_pmtct_child = qs_pmtct_child.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_child = qs_pmtct_child.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_pmtct_child = qs_pmtct_child.annotate(period=F('quarter'))
    qs_pmtct_child = qs_pmtct_child.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_pmtct_child = qs_pmtct_child.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_pmtct_child = list(val_pmtct_child)

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_child_meta, val_pmtct_child, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_child2 = list(gen_raster)

    target_de_names = (
        'HTC_TST_TARGET',
        'HTC_TST_POS_TARGET',
    )
    de_target_meta = list(product(target_de_names, subcategory_names))

    # targets are annual, so filter by year component of period and divide result by 4 to get quarter
    qs_target = DataValue.objects.what(*target_de_names).filter(year=filter_period[:4])

    qs_target = qs_target.annotate(cat_combo=F('category_combo__name'))
    qs_target = qs_target.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_target = qs_target.annotate(period=F('quarter'))
    qs_target = qs_target.order_by('district', 'subcounty', 'facility', '-de_name', 'cat_combo', 'period') # note reversed order of data element names
    val_target = qs_target.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value')/4)

    gen_raster = grabbag.rasterize(ou_list, de_target_meta, val_target, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_target2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_positivity2, val_pmtct_mother2, val_pmtct_mother_pos2, val_pmtct_child2, val_target2), key=lambda x: (x['district'], x['subcounty'], x['facility'])), key=lambda x: (x['district'], x['subcounty'], x['facility']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (district_subcounty_facility, (linked_under15_f, linked_under15_m, linked_over15_f, linked_over15_m, tst_under15_f, tst_under15_m, tst_over15_f, tst_over15_m, pos_under15_f, pos_under15_m, pos_over15_f, pos_over15_m, tst_pregnant, pos_pregnant, pos_infant, pos_pcr1, pos_pcr2, tst_male_partner, pos_male_partner, *other_vals)) = _group
        
        calculated_vals = list()

        under15_f_sum = default_zero(tst_under15_f['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_f_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_sum,
        }
        calculated_vals.append(under15_f_val)
        
        under15_m_sum = default_zero(tst_under15_m['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_m_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_sum,
        }
        calculated_vals.append(under15_m_val)
        
        over15_f_sum = default_zero(tst_over15_f['numeric_sum']) + default_zero(tst_pregnant['numeric_sum'])
        over15_f_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_sum,
        }
        calculated_vals.append(over15_f_val)
        
        over15_m_sum = default_zero(tst_over15_m['numeric_sum']) + default_zero(tst_male_partner['numeric_sum'])
        over15_m_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_sum,
        }
        calculated_vals.append(over15_m_val)
        
        half_pos_pcr = Decimal(default_zero(pos_pcr1['numeric_sum']) + default_zero(pos_pcr1['numeric_sum']))/2
        pos_under15_f_sum = default_zero(pos_under15_f['numeric_sum']) + half_pos_pcr
        pos_under15_f_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_sum,
        }
        calculated_vals.append(pos_under15_f_val)
        
        pos_under15_m_sum = default_zero(pos_under15_m['numeric_sum']) + half_pos_pcr
        pos_under15_m_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_sum,
        }
        calculated_vals.append(pos_under15_m_val)
        
        pos_over15_f_sum = default_zero(pos_over15_f['numeric_sum']) + Decimal(default_zero(pos_pregnant['numeric_sum']))
        pos_over15_f_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_sum,
        }
        calculated_vals.append(pos_over15_f_val)
        
        pos_over15_m_sum = default_zero(pos_over15_m['numeric_sum']) + Decimal(default_zero(pos_male_partner['numeric_sum']))
        pos_over15_m_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_sum,
        }
        calculated_vals.append(pos_over15_m_val)

        tested_total = sum([under15_f_sum, under15_m_sum, over15_f_sum, over15_m_sum])
        pos_total = sum([pos_under15_f_sum, pos_under15_m_sum, pos_over15_f_sum, pos_over15_m_sum])
        tested_total_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested',
            'cat_combo': None,
            'numeric_sum': tested_total,
        }
        pos_total_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+',
            'cat_combo': None,
            'numeric_sum': pos_total,
        }
        calculated_vals.append(tested_total_val)
        calculated_vals.append(pos_total_val)

        # copy linked to care totals over
        calculated_vals.append(linked_under15_f)
        calculated_vals.append(linked_under15_m)
        calculated_vals.append(linked_over15_f)
        calculated_vals.append(linked_over15_m)

        # calculate the percentages
        target_under15_f, target_under15_m, target_over15_f, target_over15_m, target_pos_under15_f, target_pos_under15_m, target_pos_over15_f, target_pos_over15_m, *further_vals = other_vals

        if all_not_none(under15_f_sum, target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            under15_f_percent = (under15_f_sum * 100) / target_under15_f['numeric_sum']
        else:
            under15_f_percent = None
        under15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_percent,
        }
        calculated_vals.append(under15_f_percent_val)

        if all_not_none(under15_m_sum, target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            under15_m_percent = (under15_m_sum * 100) / target_under15_m['numeric_sum']
        else:
            under15_m_percent = None
        under15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_percent,
        }
        calculated_vals.append(under15_m_percent_val)

        if all_not_none(over15_f_sum, target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            over15_f_percent = (over15_f_sum * 100) / target_over15_f['numeric_sum']
        else:
            over15_f_percent = None
        over15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_percent,
        }
        calculated_vals.append(over15_f_percent_val)

        if all_not_none(over15_m_sum, target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            over15_m_percent = (over15_m_sum * 100) / target_over15_m['numeric_sum']
        else:
            over15_m_percent = None
        over15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_percent,
        }
        calculated_vals.append(over15_m_percent_val)

        if all_not_none(pos_under15_f_sum, target_pos_under15_f['numeric_sum']) and target_pos_under15_f['numeric_sum']:
            pos_under15_f_percent = (pos_under15_f_sum * 100) / target_pos_under15_f['numeric_sum']
        else:
            pos_under15_f_percent = None
        pos_under15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_percent,
        }
        calculated_vals.append(pos_under15_f_percent_val)

        if all_not_none(pos_under15_m_sum, target_pos_under15_m['numeric_sum']) and target_pos_under15_m['numeric_sum']:
            pos_under15_m_percent = (pos_under15_m_sum * 100) / target_pos_under15_m['numeric_sum']
        else:
            pos_under15_m_percent = None
        pos_under15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_percent,
        }
        calculated_vals.append(pos_under15_m_percent_val)

        if all_not_none(pos_over15_f_sum, target_pos_over15_f['numeric_sum']) and target_pos_over15_f['numeric_sum']:
            pos_over15_f_percent = (pos_over15_f_sum * 100) / target_pos_over15_f['numeric_sum']
        else:
            pos_over15_f_percent = None
        pos_over15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_percent,
        }
        calculated_vals.append(pos_over15_f_percent_val)

        if all_not_none(pos_over15_m_sum, target_pos_over15_m['numeric_sum']) and target_pos_over15_m['numeric_sum']:
            pos_over15_m_percent = (pos_over15_m_sum * 100) / target_pos_over15_m['numeric_sum']
        else:
            pos_over15_m_percent = None
        pos_over15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_percent,
        }
        calculated_vals.append(pos_over15_m_percent_val)

        if all_not_none(linked_under15_f['numeric_sum'], pos_under15_f['numeric_sum']) and pos_under15_f['numeric_sum']:
            linked_under15_f_percent = (linked_under15_f['numeric_sum'] * 100) / pos_under15_f['numeric_sum']
        else:
            linked_under15_f_percent = None
        linked_under15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': linked_under15_f_percent,
        }
        calculated_vals.append(linked_under15_f_percent_val)

        if all_not_none(linked_under15_m['numeric_sum'], pos_under15_m['numeric_sum']) and pos_under15_m['numeric_sum']:
            linked_under15_m_percent = (linked_under15_m['numeric_sum'] * 100) / pos_under15_m['numeric_sum']
        else:
            linked_under15_m_percent = None
        linked_under15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': linked_under15_m_percent,
        }
        calculated_vals.append(linked_under15_m_percent_val)

        if all_not_none(linked_over15_f['numeric_sum'], pos_over15_f['numeric_sum']) and pos_over15_f['numeric_sum']:
            linked_over15_f_percent = (linked_over15_f['numeric_sum'] * 100) / pos_over15_f['numeric_sum']
        else:
            linked_over15_f_percent = None
        linked_over15_f_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': linked_over15_f_percent,
        }
        calculated_vals.append(linked_over15_f_percent_val)

        if all_not_none(linked_over15_m['numeric_sum'], pos_over15_m['numeric_sum']) and pos_over15_m['numeric_sum']:
            linked_over15_m_percent = (linked_over15_m['numeric_sum'] * 100) / pos_over15_m['numeric_sum']
        else:
            linked_over15_m_percent = None
        linked_over15_m_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': linked_over15_m_percent,
        }
        calculated_vals.append(linked_over15_m_percent_val)

        # _group[1].extend(calculated_vals)
        _group[1] = calculated_vals
    
    data_element_names = list()
    # data_element_names += list(product(hts_short_names, subcategory_names))
    # data_element_names += de_pmtct_mother_meta
    # data_element_names += de_pmtct_mother_pos_meta
    # data_element_names += list(product(pmtct_child_short_names, (None,)))
    # data_element_names += de_target_meta
    data_element_names += list(product(['Tested',], subcategory_names))
    data_element_names += list(product(['HIV+',], subcategory_names))
    data_element_names += list(product(['Tested',], [None,]))
    data_element_names += list(product(['HIV+',], [None,]))
    data_element_names += list(product(['Linked',], subcategory_names))
    data_element_names += list(product(['Tested (%)',], subcategory_names))
    data_element_names += list(product(['HIV+ (%)',], subcategory_names))
    data_element_names += list(product(['Linked (%)',], subcategory_names))

    context = {
        'grouped_data': grouped_vals,
        'val_pmtct_child': list(val_pmtct_child),
        'val_pmtct_child2': list(val_pmtct_child2),
        # 'grouped_data_size': len(grouped_vals),
        'data_element_names': data_element_names,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
    }

    return render(request, 'cannula/hts_sites.html', context)

@login_required
def hts_by_district(request):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YRS = ['%d' % (y,) for y in range(this_year, this_year-6, -1)]

    if 'period' in request.GET and request.GET['period'] in PREV_5YRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d' % (this_year,)

    period_desc = filter_period

    hts_de_names = (
        '105-4 Number of clients who have been linked to care',
        '105-4 Number of Individuals who received HIV test results',
        '105-4 Number of Individuals who tested HIV positive',
    )
    hts_short_names = (
        'Linked',
        'Tested',
        'HIV+',
    )
    subcategory_names = ['(<15, Female)', '(<15, Male)', '(15+, Female)', '(15+, Male)']
    de_positivity_meta = list(product(hts_de_names, subcategory_names))

    qs_positivity = DataValue.objects.what(*hts_de_names).filter(year=filter_period)

    cc_lt_15 = ['18 Mths-<5 Years', '5-<10 Years', '10-<15 Years']
    cc_ge_15 = ['15-<19 Years', '19-<49 Years', '>49 Years']
    #TODO: cc_lt_15_f = CategoryCombo.from_cat_names(['Female', '<15']) gives a CategoryCombo instance that makes the Case statement clearer/safer
    qs_positivity = qs_positivity.annotate(
        cat_combo=Case(
            When(Q(category_combo__categories__name__in=cc_lt_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[0])),
            When(Q(category_combo__categories__name__in=cc_lt_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[1])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & Q(category_combo__name__contains='Female'), then=Value(subcategory_names[2])),
            When(Q(category_combo__categories__name__in=cc_ge_15) & ~Q(category_combo__name__contains='Female'), then=Value(subcategory_names[3])),
            default=None, output_field=CharField()
        )
    )
    qs_positivity = qs_positivity.exclude(cat_combo__iexact=None)

    qs_positivity = qs_positivity.annotate(district=F('org_unit__parent__parent__name'))
    qs_positivity = qs_positivity.annotate(period=F('year'))
    qs_positivity = qs_positivity.order_by('district', 'de_name', 'cat_combo', 'period')
    val_positivity = qs_positivity.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_positivity = list(val_positivity)
    
    # all districts (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=1).annotate(district=F('name'))
    ou_list = list(v for v in qs_ou.values_list('district'))

    def val_with_subcat_fun(row, col):
        district, = row
        de_name, subcategory = col
        return { 'district': district, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }
    gen_raster = grabbag.rasterize(ou_list, de_positivity_meta, val_positivity, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_positivity2 = list(gen_raster)

    pmtct_mother_de_names = (
        '105-2.1 Pregnant Women newly tested for HIV this pregnancy(TR & TRR)',
        '105-2.2a Women tested for HIV in labour (1st time this Pregnancy)',
        '105-2.3a Breastfeeding mothers tested for HIV(1st test)',
    )
    de_pmtct_mother_meta = list(product(('Pregnant Women tested for HIV',), (None,)))

    qs_pmtct_mother = DataValue.objects.what(*pmtct_mother_de_names).filter(year=filter_period)
    qs_pmtct_mother = qs_pmtct_mother.annotate(de_name=Value('Pregnant Women tested for HIV', output_field=CharField()))
    qs_pmtct_mother = qs_pmtct_mother.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother = qs_pmtct_mother.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_mother = qs_pmtct_mother.annotate(period=F('year'))
    qs_pmtct_mother = qs_pmtct_mother.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother = qs_pmtct_mother.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_meta, val_pmtct_mother, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother2 = list(gen_raster)

    pmtct_mother_pos_de_names = (
        '105-2.1 A19:Pregnant Women testing HIV+ on a retest (TRR+)',
        '105-2.2a Women testing HIV+ in labour (1st time this Pregnancy)',
        '105-2.2b Women testing HIV+ in labour (Retest this Pregnancy)',
        '105-2.3a Breastfeeding mothers newly testing HIV+(1st test)',
        '105-2.3b Breastfeeding mothers newly testing HIV+(retest)',
    )
    de_pmtct_mother_pos_meta = list(product(('Pregnant Women testing HIV+',), (None,)))

    qs_pmtct_mother_pos = DataValue.objects.what(*pmtct_mother_pos_de_names).filter(year=filter_period)
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(de_name=Value('Pregnant Women testing HIV+', output_field=CharField()))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.annotate(period=F('year'))
    qs_pmtct_mother_pos = qs_pmtct_mother_pos.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_mother_pos = qs_pmtct_mother_pos.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_mother_pos_meta, val_pmtct_mother_pos, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_mother_pos2 = list(gen_raster)

    pmtct_child_de_names = (
        '105-2.4a Exposed Infants Tested for HIV Below 18 Months(by 1st PCR) ',
        '105-2.4b 1st DNA PCR result returned(HIV+)',
        '105-2.4b 2nd DNA PCR result returned(HIV+)',
        '105-2.1a Male partners received HIV test results in eMTCT(Total)',
        '105-2.1b Male partners received HIV test results in eMTCT(HIV+)',
    )
    pmtct_child_short_names = (
        'PMTCT INFANT HIV+',
        'PMTCT CHILD PCR1 HIV+',
        'PMTCT CHILD PCR2 HIV+',
        'PMTCT MALE PARTNERS TESTED',
        'PMTCT MALE PARTNERS HIV+',
    )
    de_pmtct_child_meta = list(product(pmtct_child_de_names, (None,)))

    qs_pmtct_child = DataValue.objects.what(*pmtct_child_de_names).filter(year=filter_period)
    qs_pmtct_child = qs_pmtct_child.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_pmtct_child = qs_pmtct_child.annotate(district=F('org_unit__parent__parent__name'))
    qs_pmtct_child = qs_pmtct_child.annotate(period=F('year'))
    qs_pmtct_child = qs_pmtct_child.order_by('district', 'de_name', 'cat_combo', 'period')
    val_pmtct_child = qs_pmtct_child.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_pmtct_child_meta, val_pmtct_child, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_pmtct_child2 = list(gen_raster)

    target_de_names = (
        'HTC_TST_TARGET',
        'HTC_TST_POS_TARGET',
    )
    de_target_meta = list(product(target_de_names, subcategory_names))

    # targets are annual, so filter by year component of period
    qs_target = DataValue.objects.what(*target_de_names).filter(year=filter_period[:4])

    qs_target = qs_target.annotate(cat_combo=F('category_combo__name'))
    qs_target = qs_target.annotate(district=F('org_unit__parent__parent__name'))
    qs_target = qs_target.annotate(period=F('year'))
    qs_target = qs_target.order_by('district', '-de_name', 'cat_combo', 'period') # note reversed order of data element names
    val_target = qs_target.values('district', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_target = list(val_target)

    gen_raster = grabbag.rasterize(ou_list, de_target_meta, val_target, lambda x: (x['district'],), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_target2 = list(gen_raster)

    # combine the data and group by district
    grouped_vals = groupbylist(sorted(chain(val_positivity2, val_pmtct_mother2, val_pmtct_mother_pos2, val_pmtct_child2, val_target2), key=lambda x: (x['district'],)), key=lambda x: (x['district'],))

    # perform calculations
    for _group in grouped_vals:
        (ou_path_list, (linked_under15_f, linked_under15_m, linked_over15_f, linked_over15_m, tst_under15_f, tst_under15_m, tst_over15_f, tst_over15_m, pos_under15_f, pos_under15_m, pos_over15_f, pos_over15_m, tst_pregnant, pos_pregnant, pos_infant, pos_pcr1, pos_pcr2, tst_male_partner, pos_male_partner, *other_vals)) = _group
        
        calculated_vals = list()

        under15_f_sum = default_zero(tst_under15_f['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_sum,
        }
        calculated_vals.append(under15_f_val)
        
        under15_m_sum = default_zero(tst_under15_m['numeric_sum']) + Decimal(default_zero(pos_infant['numeric_sum'])/2)
        under15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_sum,
        }
        calculated_vals.append(under15_m_val)
        
        over15_f_sum = default_zero(tst_over15_f['numeric_sum']) + default_zero(tst_pregnant['numeric_sum'])
        over15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_sum,
        }
        calculated_vals.append(over15_f_val)
        
        over15_m_sum = default_zero(tst_over15_m['numeric_sum']) + default_zero(tst_male_partner['numeric_sum'])
        over15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_sum,
        }
        calculated_vals.append(over15_m_val)
        
        half_pos_pcr = Decimal(default_zero(pos_pcr1['numeric_sum']) + default_zero(pos_pcr1['numeric_sum']))/2
        pos_under15_f_sum = default_zero(pos_under15_f['numeric_sum']) + half_pos_pcr
        pos_under15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_sum,
        }
        calculated_vals.append(pos_under15_f_val)
        
        pos_under15_m_sum = default_zero(pos_under15_m['numeric_sum']) + half_pos_pcr
        pos_under15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_sum,
        }
        calculated_vals.append(pos_under15_m_val)
        
        pos_over15_f_sum = default_zero(pos_over15_f['numeric_sum']) + Decimal(default_zero(pos_pregnant['numeric_sum']))
        pos_over15_f_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_sum,
        }
        calculated_vals.append(pos_over15_f_val)
        
        pos_over15_m_sum = default_zero(pos_over15_m['numeric_sum']) + Decimal(default_zero(pos_male_partner['numeric_sum']))
        pos_over15_m_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_sum,
        }
        calculated_vals.append(pos_over15_m_val)

        tested_total = sum([under15_f_sum, under15_m_sum, over15_f_sum, over15_m_sum])
        pos_total = sum([pos_under15_f_sum, pos_under15_m_sum, pos_over15_f_sum, pos_over15_m_sum])
        tested_total_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested',
            'cat_combo': None,
            'numeric_sum': tested_total,
        }
        pos_total_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+',
            'cat_combo': None,
            'numeric_sum': pos_total,
        }
        calculated_vals.append(tested_total_val)
        calculated_vals.append(pos_total_val)

        # copy linked to care totals over
        calculated_vals.append(linked_under15_f)
        calculated_vals.append(linked_under15_m)
        calculated_vals.append(linked_over15_f)
        calculated_vals.append(linked_over15_m)

        # calculate the percentages
        target_under15_f, target_under15_m, target_over15_f, target_over15_m, target_pos_under15_f, target_pos_under15_m, target_pos_over15_f, target_pos_over15_m, *further_vals = other_vals

        if all_not_none(under15_f_sum, target_under15_f['numeric_sum']) and target_under15_f['numeric_sum']:
            under15_f_percent = (under15_f_sum * 100) / target_under15_f['numeric_sum']
        else:
            under15_f_percent = None
        under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': under15_f_percent,
        }
        calculated_vals.append(under15_f_percent_val)

        if all_not_none(under15_m_sum, target_under15_m['numeric_sum']) and target_under15_m['numeric_sum']:
            under15_m_percent = (under15_m_sum * 100) / target_under15_m['numeric_sum']
        else:
            under15_m_percent = None
        under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': under15_m_percent,
        }
        calculated_vals.append(under15_m_percent_val)

        if all_not_none(over15_f_sum, target_over15_f['numeric_sum']) and target_over15_f['numeric_sum']:
            over15_f_percent = (over15_f_sum * 100) / target_over15_f['numeric_sum']
        else:
            over15_f_percent = None
        over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': over15_f_percent,
        }
        calculated_vals.append(over15_f_percent_val)

        if all_not_none(over15_m_sum, target_over15_m['numeric_sum']) and target_over15_m['numeric_sum']:
            over15_m_percent = (over15_m_sum * 100) / target_over15_m['numeric_sum']
        else:
            over15_m_percent = None
        over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Tested (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': over15_m_percent,
        }
        calculated_vals.append(over15_m_percent_val)

        if all_not_none(pos_under15_f_sum, target_pos_under15_f['numeric_sum']) and target_pos_under15_f['numeric_sum']:
            pos_under15_f_percent = (pos_under15_f_sum * 100) / target_pos_under15_f['numeric_sum']
        else:
            pos_under15_f_percent = None
        pos_under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': pos_under15_f_percent,
        }
        calculated_vals.append(pos_under15_f_percent_val)

        if all_not_none(pos_under15_m_sum, target_pos_under15_m['numeric_sum']) and target_pos_under15_m['numeric_sum']:
            pos_under15_m_percent = (pos_under15_m_sum * 100) / target_pos_under15_m['numeric_sum']
        else:
            pos_under15_m_percent = None
        pos_under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': pos_under15_m_percent,
        }
        calculated_vals.append(pos_under15_m_percent_val)

        if all_not_none(pos_over15_f_sum, target_pos_over15_f['numeric_sum']) and target_pos_over15_f['numeric_sum']:
            pos_over15_f_percent = (pos_over15_f_sum * 100) / target_pos_over15_f['numeric_sum']
        else:
            pos_over15_f_percent = None
        pos_over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': pos_over15_f_percent,
        }
        calculated_vals.append(pos_over15_f_percent_val)

        if all_not_none(pos_over15_m_sum, target_pos_over15_m['numeric_sum']) and target_pos_over15_m['numeric_sum']:
            pos_over15_m_percent = (pos_over15_m_sum * 100) / target_pos_over15_m['numeric_sum']
        else:
            pos_over15_m_percent = None
        pos_over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'HIV+ (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': pos_over15_m_percent,
        }
        calculated_vals.append(pos_over15_m_percent_val)

        if all_not_none(linked_under15_f['numeric_sum'], pos_under15_f['numeric_sum']) and pos_under15_f['numeric_sum']:
            linked_under15_f_percent = (linked_under15_f['numeric_sum'] * 100) / pos_under15_f['numeric_sum']
        else:
            linked_under15_f_percent = None
        linked_under15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Female)',
            'numeric_sum': linked_under15_f_percent,
        }
        calculated_vals.append(linked_under15_f_percent_val)

        if all_not_none(linked_under15_m['numeric_sum'], pos_under15_m['numeric_sum']) and pos_under15_m['numeric_sum']:
            linked_under15_m_percent = (linked_under15_m['numeric_sum'] * 100) / pos_under15_m['numeric_sum']
        else:
            linked_under15_m_percent = None
        linked_under15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(<15, Male)',
            'numeric_sum': linked_under15_m_percent,
        }
        calculated_vals.append(linked_under15_m_percent_val)

        if all_not_none(linked_over15_f['numeric_sum'], pos_over15_f['numeric_sum']) and pos_over15_f['numeric_sum']:
            linked_over15_f_percent = (linked_over15_f['numeric_sum'] * 100) / pos_over15_f['numeric_sum']
        else:
            linked_over15_f_percent = None
        linked_over15_f_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Female)',
            'numeric_sum': linked_over15_f_percent,
        }
        calculated_vals.append(linked_over15_f_percent_val)

        if all_not_none(linked_over15_m['numeric_sum'], pos_over15_m['numeric_sum']) and pos_over15_m['numeric_sum']:
            linked_over15_m_percent = (linked_over15_m['numeric_sum'] * 100) / pos_over15_m['numeric_sum']
        else:
            linked_over15_m_percent = None
        linked_over15_m_percent_val = {
            'district': ou_path_list[0],
            'de_name': 'Linked (%)',
            'cat_combo': '(15+, Male)',
            'numeric_sum': linked_over15_m_percent,
        }
        calculated_vals.append(linked_over15_m_percent_val)

        # _group[1].extend(calculated_vals)
        _group[1] = calculated_vals
    
    data_element_names = list()
    
    # data_element_names += list(product(hts_short_names, subcategory_names))
    # data_element_names += de_pmtct_mother_meta
    # data_element_names += de_pmtct_mother_pos_meta
    # data_element_names += list(product(pmtct_child_short_names, (None,)))
    # data_element_names += de_target_meta

    data_element_names += list(product(['Tested',], subcategory_names))
    data_element_names += list(product(['HIV+',], subcategory_names))
    data_element_names += list(product(['Tested',], [None,]))
    data_element_names += list(product(['HIV+',], [None,]))
    data_element_names += list(product(['Linked',], subcategory_names))
    data_element_names += list(product(['Tested (%)',], subcategory_names))
    data_element_names += list(product(['HIV+ (%)',], subcategory_names))
    data_element_names += list(product(['Linked (%)',], subcategory_names))

    context = {
        'grouped_data': grouped_vals,
        'ou_list': ou_list,
        'val_target': val_target,
        'val_target2': val_target2,
        # 'grouped_data_size': len(grouped_vals),
        'data_element_names': data_element_names,
        'period_desc': period_desc,
        'period_list': PREV_5YRS,
    }

    return render(request, 'cannula/hts_districts.html', context)

@login_required
def vmmc_by_site(request):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=3).annotate(district=F('parent__parent__name'), subcounty=F('parent__name'), facility=F('name'))
    ou_list = list(qs_ou.values_list('district', 'subcounty', 'facility'))

    def val_with_subcat_fun(row, col):
        district, subcounty, facility = row
        de_name, subcategory = col
        return { 'district': district, 'subcounty': subcounty, 'facility': facility, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }

    targets_de_names = (
        'VMMC_CIRC_TARGET',
        'VMMC_DEVICE_TARGET',
        'VMMC_SURGICAL_TARGET',
    )
    targets_short_names = (
        'TARGET: VMMC_CIRC',
        'TARGET: Device-based',
        'TARGET: Surgical',
    )
    de_targets_meta = list(product(targets_de_names, (None,)))

    qs_targets = DataValue.objects.what(*targets_de_names).filter(quarter=filter_period)
    qs_targets = qs_targets.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_targets = qs_targets.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_targets = qs_targets.annotate(period=F('quarter'))
    qs_targets = qs_targets.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_targets = qs_targets.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_targets = list(val_targets)

    gen_raster = grabbag.rasterize(ou_list, de_targets_meta, val_targets, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_targets2 = list(gen_raster)

    method_de_names = (
        '105-5 Clients circumcised by circumcision Technique Device Based (DC)',
        '105-5 Clients circumcised by circumcision Technique Other VMMC techniques',
        '105-5 Clients circumcised by circumcision Technique Surgical(SC)',
    )
    method_short_names = (
        'Circumcised by technique - Device Based',
        'Circumcised by technique - Other',
        'Circumcised by technique - Surgical',
    )
    de_method_meta = list(product(method_de_names, (None,)))

    qs_method = DataValue.objects.what(*method_de_names).filter(quarter=filter_period)
    qs_method = qs_method.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_method = qs_method.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_method = qs_method.annotate(period=F('quarter'))
    qs_method = qs_method.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_method = qs_method.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_method_meta, val_method, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_method2 = list(gen_raster)

    hiv_de_names = (
        '105-5 SMC Clients Counseled, Tested and Circumcised for HIV at SMC site HIV Negative',
        '105-5 SMC Clients Counseled, Tested and Circumcised for HIV at SMC site HIV Positive',
    )
    hiv_short_names = (
        'Circumcised by HIV status - Negative',
        'Circumcised by HIV status - Positive',
    )
    de_hiv_meta = list(product(hiv_de_names, (None,)))

    qs_hiv = DataValue.objects.what(*hiv_de_names).filter(quarter=filter_period)
    qs_hiv = qs_hiv.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_hiv = qs_hiv.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_hiv = qs_hiv.annotate(period=F('quarter'))
    qs_hiv = qs_hiv.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_hiv = qs_hiv.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_hiv_meta, val_hiv, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_hiv2 = list(gen_raster)

    location_de_names = (
        '105-5 Number of Males Circumcised by Age group and Technique Facility, Device Based (DC)',
        '105-5 Number of Males Circumcised by Age group and Technique Facility, Surgical(SC)',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach, Device Based (DC)',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach, Surgical(SC)',
    )
    location_de_names2 = (
        '105-5 Number of Males Circumcised by Age group and Technique Facility',
        '105-5 Number of Males Circumcised by Age group and Technique Outreach',
    )
    location_prefix_len = len('105-5 Number of Males Circumcised by Age group and Technique Facility')
    location_short_names = (
        'Circumcised by site type - Static',
        'Circumcised by site type - Mobile',
    )
    de_location_meta = list(product(location_de_names2, (None,)))

    qs_location = DataValue.objects.what(*location_de_names).filter(quarter=filter_period)
    qs_location = qs_location.annotate(cat_combo=Value(None, output_field=CharField()))

    # drop the technique section from the returned data element name
    qs_location = qs_location.annotate(de_name=Substr('data_element__name', 1, location_prefix_len))

    qs_location = qs_location.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_location = qs_location.annotate(period=F('quarter'))
    qs_location = qs_location.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_location = qs_location.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_location_meta, val_location, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_location2 = list(gen_raster)

    followup_de_names = (
        '105-5a Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 48 Hours)',
        '105-5b Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Within 7 Days)',
        '105-5c Number of Clients Circumcised who Returned for Follow Up Visit within 6 weeks of SMC Procedure(Beyond 7 Days)',
    )
    followup_short_names = (
        'Follow up - Within 48 hours',
        'Follow up - Within 7 days',
        'Follow up - Beyond 7 days',
    )
    de_followup_meta = list(product(followup_de_names, (None,)))

    qs_followup = DataValue.objects.what(*followup_de_names).filter(quarter=filter_period)
    qs_followup = qs_followup.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_followup = qs_followup.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_followup = qs_followup.annotate(period=F('quarter'))
    qs_followup = qs_followup.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_followup = qs_followup.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_followup_meta, val_followup, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_followup2 = list(gen_raster)

    adverse_de_names = (
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Moderate',
        '105-5 Clients Circumcised who Experienced one or more Adverse Events Severe',
    )
    adverse_short_names = (
        'Adverse Events - Moderate',
        'Adverse Events - Severe',
    )
    de_adverse_meta = list(product(adverse_de_names, (None,)))

    qs_adverse = DataValue.objects.what(*adverse_de_names).filter(quarter=filter_period)
    qs_adverse = qs_adverse.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_adverse = qs_adverse.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_adverse = qs_adverse.annotate(period=F('quarter'))
    qs_adverse = qs_adverse.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_adverse = qs_adverse.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))

    gen_raster = grabbag.rasterize(ou_list, de_adverse_meta, val_adverse, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_adverse2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_targets2, val_hiv2, val_location2, val_method2, val_followup2, val_adverse2), key=lambda x: (x['district'], x['subcounty'], x['facility'])), key=lambda x: (x['district'], x['subcounty'], x['facility']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (district_subcounty_facility, (target_total, target_device, target_surgical, hiv_negative, hiv_positive, location_facility, location_outreach, method_device, method_other, method_surgical, followup_48hrs, followup_7days, followup_plus7days, adverse_moderate, adverse_severe, *other_vals)) = _group
        
        calculated_vals = list()

        method_sum = default_zero(method_device['numeric_sum']) + default_zero(method_surgical['numeric_sum']) + default_zero(method_other['numeric_sum'])

        if all_not_none(target_total['numeric_sum'], method_sum) and target_total['numeric_sum']:
            target_total_percent = (method_sum * 100) / target_total['numeric_sum']
        else:
            target_total_percent = None
        target_total_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Perf% Circumcised',
            'cat_combo': None,
            'numeric_sum': target_total_percent,
        }
        calculated_vals.append(target_total_percent_val)

        if all_not_none(target_device['numeric_sum'], method_device['numeric_sum']) and target_device['numeric_sum']:
            if target_device['numeric_sum'] == 4:
                tt = 1/0
            target_device_percent = (method_device['numeric_sum'] * 100) / target_device['numeric_sum']
        else:
            target_device_percent = None
        target_device_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Perf% Circumcised DC',
            'cat_combo': None,
            'numeric_sum': target_device_percent,
        }
        calculated_vals.append(target_device_percent_val)

        if all_not_none(target_surgical['numeric_sum'], method_surgical['numeric_sum']) and target_surgical['numeric_sum']:
            target_surgical_percent = (method_surgical['numeric_sum'] * 100) / target_surgical['numeric_sum']
        else:
            target_surgical_percent = None
        target_surgical_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Perf% Circumcised Surgical',
            'cat_combo': None,
            'numeric_sum': target_surgical_percent,
        }
        calculated_vals.append(target_surgical_percent_val)

        if all_not_none(followup_48hrs['numeric_sum'], method_sum) and method_sum:
            followup_48hrs_percent = (followup_48hrs['numeric_sum'] * 100) / method_sum
        else:
            followup_48hrs_percent = None
        followup_48hrs_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': '% who returned within 48 hours',
            'cat_combo': None,
            'numeric_sum': followup_48hrs_percent,
        }
        calculated_vals.append(followup_48hrs_percent_val)

        adverse_sum = default_zero(adverse_moderate['numeric_sum']) + default_zero(adverse_severe['numeric_sum'])

        if all_not_none(adverse_sum, method_sum) and method_sum:
            adverse_percent = (adverse_sum * 100) / method_sum
        else:
            adverse_percent = None
        adverse_percent_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': '% with at least one adverse event',
            'cat_combo': None,
            'numeric_sum': adverse_percent,
        }
        calculated_vals.append(adverse_percent_val)

        _group[1].extend(calculated_vals)

    data_element_names = list()
    data_element_names += list(product(targets_short_names, (None,)))
    data_element_names += list(product(hiv_short_names, (None,)))
    data_element_names += list(product(location_short_names, (None,)))
    data_element_names += list(product(method_short_names, (None,)))
    data_element_names += list(product(followup_short_names, (None,)))
    data_element_names += list(product(adverse_short_names, (None,)))

    data_element_names += list(product(['Perf% Circumcised'], (None,)))
    data_element_names += list(product(['Perf% Circumcised DC'], (None,)))
    data_element_names += list(product(['Perf% Circumcised Surgical'], (None,)))
    data_element_names += list(product(['% who returned within 48 hours'], (None,)))
    data_element_names += list(product(['% with at least one adverse event'], (None,)))

    context = {
        'grouped_data': grouped_vals,
        'ou_list': ou_list,
        'val_targets': val_targets,
        'val_targets2': val_targets2,
        'data_element_names': data_element_names,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
    }

    return render(request, 'cannula/vmmc_sites.html', context)

@login_required
def lab_by_site(request):
    this_day = date.today()
    this_year = this_day.year
    PREV_5YR_QTRS = ['%d-Q%d' % (y, q) for y in range(this_year, this_year-6, -1) for q in range(4, 0, -1)]

    if 'period' in request.GET and request.GET['period'] in PREV_5YR_QTRS:
        filter_period=request.GET['period']
    else:
        filter_period = '%d-Q%d' % (this_year, month2quarter(this_day.month))

    period_desc = dateutil.DateSpan.fromquarter(filter_period).format()

    # # all facilities (or equivalent)
    qs_ou = OrgUnit.objects.filter(level=3).annotate(district=F('parent__parent__name'), subcounty=F('parent__name'), facility=F('name'))
    ou_list = list(qs_ou.values_list('district', 'subcounty', 'facility'))

    def val_with_subcat_fun(row, col):
        district, subcounty, facility = row
        de_name, subcategory = col
        return { 'district': district, 'subcounty': subcounty, 'facility': facility, 'cat_combo': subcategory, 'de_name': de_name, 'numeric_sum': None }

    malaria_de_names = (
        '105-7.3 Lab Malaria Microscopy  Number Done',
        '105-7.3 Lab Malaria RDTs Number Done',
    )
    malaria_short_names = (
        'Malaria Microscopy Done',
        'Malaria RDTs Done',
    )
    de_malaria_meta = list(product(malaria_de_names, (None,)))

    qs_malaria = DataValue.objects.what(*malaria_de_names).filter(quarter=filter_period)
    qs_malaria = qs_malaria.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_malaria = qs_malaria.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_malaria = qs_malaria.annotate(period=F('quarter'))
    qs_malaria = qs_malaria.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_malaria = qs_malaria.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_malaria = list(val_malaria)

    gen_raster = grabbag.rasterize(ou_list, de_malaria_meta, val_malaria, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_malaria2 = list(gen_raster)

    hiv_determine_de_names = (
        '105-7.8 Lab Determine Clinical Diagnosis',
        '105-7.8 Lab Determine HCT',
        '105-7.8 Lab Determine PMTCT',
        '105-7.8 Lab Determine Quality Control',
        '105-7.8 Lab Determine SMC',
    )
    hiv_determine_short_names = (
        'HIV tests done using Determine',
    )
    de_hiv_determine_meta = list(product(['HIV tests done using Determine'], (None,)))

    qs_hiv_determine = DataValue.objects.what(*hiv_determine_de_names).filter(quarter=filter_period)
    qs_hiv_determine = qs_hiv_determine.annotate(de_name=Value('HIV tests done using Determine', output_field=CharField()))
    qs_hiv_determine = qs_hiv_determine.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_hiv_determine = qs_hiv_determine.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_hiv_determine = qs_hiv_determine.annotate(period=F('quarter'))
    qs_hiv_determine = qs_hiv_determine.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_hiv_determine = qs_hiv_determine.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_determine = list(val_hiv_determine)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_determine_meta, val_hiv_determine, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_hiv_determine2 = list(gen_raster)

    hiv_statpak_de_names = (
        '105-7.8 Lab Stat pak  Clinical Diagnosis',
        '105-7.8 Lab Stat pak  HCT',
        '105-7.8 Lab Stat pak  PMTCT',
        '105-7.8 Lab Stat pak  Quality Control',
        '105-7.8 Lab Stat pak  SMC',
    )
    hiv_statpak_short_names = (
        'HIV tests done using Stat Pak',
    )
    de_hiv_statpak_meta = list(product(['HIV tests done using Stat Pak'], (None,)))

    qs_hiv_statpak = DataValue.objects.what(*hiv_statpak_de_names).filter(quarter=filter_period)
    qs_hiv_statpak = qs_hiv_statpak.annotate(de_name=Value('HIV tests done using Stat Pak', output_field=CharField()))
    qs_hiv_statpak = qs_hiv_statpak.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_hiv_statpak = qs_hiv_statpak.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_hiv_statpak = qs_hiv_statpak.annotate(period=F('quarter'))
    qs_hiv_statpak = qs_hiv_statpak.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_hiv_statpak = qs_hiv_statpak.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_statpak = list(val_hiv_statpak)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_statpak_meta, val_hiv_statpak, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_hiv_statpak2 = list(gen_raster)

    hiv_unigold_de_names = (
        '105-7.8 Lab Unigold Clinical Diagnosis',
        '105-7.8 Lab Unigold HCT',
        '105-7.8 Lab Unigold PMTCT',
        '105-7.8 Lab Unigold Quality Control',
        '105-7.8 Lab Unigold SMC',
    )
    hiv_unigold_short_names = (
        'HIV tests done using Unigold',
    )
    de_hiv_unigold_meta = list(product(['HIV tests done using Unigold'], (None,)))

    qs_hiv_unigold = DataValue.objects.what(*hiv_unigold_de_names).filter(quarter=filter_period)
    qs_hiv_unigold = qs_hiv_unigold.annotate(de_name=Value('HIV tests done using Unigold', output_field=CharField()))
    qs_hiv_unigold = qs_hiv_unigold.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_hiv_unigold = qs_hiv_unigold.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_hiv_unigold = qs_hiv_unigold.annotate(period=F('quarter'))
    qs_hiv_unigold = qs_hiv_unigold.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_hiv_unigold = qs_hiv_unigold.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_hiv_unigold = list(val_hiv_unigold)

    gen_raster = grabbag.rasterize(ou_list, de_hiv_unigold_meta, val_hiv_unigold, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_hiv_unigold2 = list(gen_raster)

    tb_smear_de_names = (
        '105-7.6 Lab ZN for AFBs  Number Done',
    )
    tb_smear_short_names = (
        'TB Smear',
    )
    de_tb_smear_meta = list(product(tb_smear_de_names, (None,)))

    qs_tb_smear = DataValue.objects.what(*tb_smear_de_names).filter(quarter=filter_period)
    qs_tb_smear = qs_tb_smear.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_tb_smear = qs_tb_smear.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_tb_smear = qs_tb_smear.annotate(period=F('quarter'))
    qs_tb_smear = qs_tb_smear.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_tb_smear = qs_tb_smear.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_tb_smear = list(val_tb_smear)

    gen_raster = grabbag.rasterize(ou_list, de_tb_smear_meta, val_tb_smear, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_tb_smear2 = list(gen_raster)

    syphilis_de_names = (
        '105-7.4 Lab VDRL/RPR Number Done',
        '105-7.4 Lab TPHA  Number Done',
    )
    syphilis_short_names = (
        'Syphilis tests',
    )
    de_syphilis_meta = list(product(['Syphilis tests'], (None,)))

    qs_syphilis = DataValue.objects.what(*syphilis_de_names).filter(quarter=filter_period)
    qs_syphilis = qs_syphilis.annotate(de_name=Value('Syphilis tests', output_field=CharField()))
    qs_syphilis = qs_syphilis.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_syphilis = qs_syphilis.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_syphilis = qs_syphilis.annotate(period=F('quarter'))
    qs_syphilis = qs_syphilis.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_syphilis = qs_syphilis.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_syphilis = list(val_syphilis)

    gen_raster = grabbag.rasterize(ou_list, de_syphilis_meta, val_syphilis, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_syphilis2 = list(gen_raster)

    liver_de_names = (
        '105-7.7 Lab ALT Number Done',
        '105-7.7 Lab AST Number Done',
        '105-7.7 Lab Albumin  Number Done',
    )
    liver_short_names = (
        'LFTs',
    )
    de_liver_meta = list(product(['LFTs'], (None,)))

    qs_liver = DataValue.objects.what(*liver_de_names).filter(quarter=filter_period)
    qs_liver = qs_liver.annotate(de_name=Value('LFTs', output_field=CharField()))
    qs_liver = qs_liver.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_liver = qs_liver.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_liver = qs_liver.annotate(period=F('quarter'))
    qs_liver = qs_liver.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_liver = qs_liver.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_liver = list(val_liver)

    gen_raster = grabbag.rasterize(ou_list, de_liver_meta, val_liver, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_liver2 = list(gen_raster)

    renal_de_names = (
        '105-7.7 Lab Calcium  Number Done',
        '105-7.7 Lab Creatinine Number Done',
        '105-7.7 Lab Potassium Number Done',
        '105-7.7 Lab Sodium Number Done',
        '105-7.7 Lab Total Protein Number Done',
        '105-7.7 Lab Urea Number Done',
    )
    renal_short_names = (
        'RFTs',
    )
    de_renal_meta = list(product(['RFTs'], (None,)))

    qs_renal = DataValue.objects.what(*renal_de_names).filter(quarter=filter_period)
    qs_renal = qs_renal.annotate(de_name=Value('RFTs', output_field=CharField()))
    qs_renal = qs_renal.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_renal = qs_renal.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_renal = qs_renal.annotate(period=F('quarter'))
    qs_renal = qs_renal.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_renal = qs_renal.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_renal = list(val_renal)

    gen_raster = grabbag.rasterize(ou_list, de_renal_meta, val_renal, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_renal2 = list(gen_raster)

    other_haem_de_names = (
        'All Other Haematology - Lab - OPD  Number Done',
    )
    other_haem_short_names = (
        'All other Haematology',
    )
    de_other_haem_meta = list(product(other_haem_de_names, (None,)))

    qs_other_haem = DataValue.objects.what(*other_haem_de_names).filter(quarter=filter_period)
    qs_other_haem = qs_other_haem.annotate(cat_combo=Value(None, output_field=CharField()))

    qs_other_haem = qs_other_haem.annotate(district=F('org_unit__parent__parent__name'), subcounty=F('org_unit__parent__name'), facility=F('org_unit__name'))
    qs_other_haem = qs_other_haem.annotate(period=F('quarter'))
    qs_other_haem = qs_other_haem.order_by('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period')
    val_other_haem = qs_other_haem.values('district', 'subcounty', 'facility', 'de_name', 'cat_combo', 'period').annotate(values_count=Count('numeric_value'), numeric_sum=Sum('numeric_value'))
    val_other_haem = list(val_other_haem)

    gen_raster = grabbag.rasterize(ou_list, de_other_haem_meta, val_other_haem, lambda x: (x['district'], x['subcounty'], x['facility']), lambda x: (x['de_name'], x['cat_combo']), val_with_subcat_fun)
    val_other_haem2 = list(gen_raster)

    # combine the data and group by district, subcounty and facility
    grouped_vals = groupbylist(sorted(chain(val_malaria2, val_hiv_determine2, val_hiv_statpak2, val_hiv_unigold2, val_tb_smear2, val_syphilis2, val_liver2,val_renal2, val_other_haem2), key=lambda x: (x['district'], x['subcounty'], x['facility'])), key=lambda x: (x['district'], x['subcounty'], x['facility']))
    if True:
        grouped_vals = list(filter_empty_rows(grouped_vals))

    # perform calculations
    for _group in grouped_vals:
        (district_subcounty_facility, (malaria_microscopy, malaria_rdt, *other_vals)) = _group
        
        calculated_vals = list()

        malaria_sum = default_zero(malaria_microscopy['numeric_sum']) + default_zero(malaria_rdt['numeric_sum'])
        malaria_val = {
            'district': district_subcounty_facility[0],
            'subcounty': district_subcounty_facility[1],
            'facility': district_subcounty_facility[2],
            'de_name': 'Malaria (Smear & RDTs)',
            'cat_combo': None,
            'numeric_sum': malaria_sum,
        }
        calculated_vals.append(malaria_val)

        _group[1].extend(calculated_vals)

    data_element_names = list()
    data_element_names += list(product(malaria_short_names, (None,)))
    data_element_names += list(product(hiv_determine_short_names, (None,)))
    data_element_names += list(product(hiv_statpak_short_names, (None,)))
    data_element_names += list(product(hiv_unigold_short_names, (None,)))
    data_element_names += list(product(tb_smear_short_names, (None,)))
    data_element_names += list(product(syphilis_short_names, (None,)))
    data_element_names += list(product(liver_short_names, (None,)))
    data_element_names += list(product(renal_short_names, (None,)))
    data_element_names += list(product(other_haem_short_names, (None,)))

    data_element_names += list(product(['Malaria (Smear & RDTs)'], (None,)))

    context = {
        'grouped_data': grouped_vals,
        'ou_list': ou_list,
        'data_element_names': data_element_names,
        'period_desc': period_desc,
        'period_list': PREV_5YR_QTRS,
    }

    return render(request, 'cannula/lab_sites.html', context)
