from django.db import models
from django.db.models import Avg, Case, Count, F, Max, Min, Prefetch, Q, Sum, When
from django.db.models.signals import post_init
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ValidationError
from django.conf import settings

import logging
logger = logging.getLogger(__name__)

import mimetypes
from functools import lru_cache, partial
from decimal import Decimal

from mptt.models import MPTTModel, TreeForeignKey

from . import grabbag

def make_random_filename(instance, filename):
    mt = mimetypes.guess_type(filename)
    file_ext = mimetypes.guess_extension(mt[0])

    return grabbag.make_random_code(code_length=16) + file_ext

fs = FileSystemStorage(location=settings.SOURCE_DOC_DIR)

class SourceDocument(models.Model):
    orig_filename = models.CharField(max_length=128, blank=True, null=True)
    file = models.FileField(upload_to=make_random_filename, storage=fs)
    uploaded_at = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        # store the original filename away for later
        self.orig_filename = self.file.name
        super(SourceDocument, self).save(*args, **kwargs)

    def __str__(self):
        return '%s: %s' % (self.file, self.orig_filename)

class OrgUnit(MPTTModel):
    name = models.CharField(max_length=64)
    parent = TreeForeignKey('self', null=True, blank=True, related_name='children', db_index=True)

    class MPTTMeta:
        order_insertion_by = ['name']

    class Meta:
        unique_together = (('name', 'parent'),)
        verbose_name = 'organisation unit'

    @classmethod
    def from_path_str(cls, path, path_sep='/'):
        return cls.from_path(path.split(path_sep))

    @classmethod
    def from_path(cls, *path_parts):
        current_node = None
        for p in path_parts:
            if p:
                ou_p, created = cls.objects.get_or_create(name=str(p), parent=current_node)
                current_node = ou_p
            else:
                break # stop processing when you find blank/empty path component/name

        return current_node

    @classmethod
    @lru_cache(maxsize=None)
    def from_path_recurse(cls, *path_parts):
        if len(path_parts) == 0:
            return None
        *parent_path, node_name = path_parts
        if len(parent_path) == 0:
            ou, created = cls.objects.get_or_create(name=node_name, parent=None)
        else:
            ou_parent = cls.from_path_recurse(*parent_path)
            ou, created = cls.objects.get_or_create(name=node_name, parent=ou_parent)
        return ou

    def __str__(self):
        return '%s [parent_id: %s]' % (self.name, str(self.parent_id),)

class DataElement(models.Model):
    VALUE_TYPES = (
        ('NUMBER', 'Number'),
        ('INTEGER', 'Integer (whole numbers only)'),
        ('POS_INT', 'Positive Integer'),
        # ('PERCENT', 'Percentage'), # implies the Average aggregation method
        #TODO: for boolean types auto-create a (hidden?) category with two options corresponding to the labels we have in the boolean
        # ('BOOLEAN', 'Boolean (True/False)'), # two values (0/1) with default labels (category options?) of 'True'/'False'
        # ('CHOICE', 'Selection from fixed set') # implies the Count aggregation method
    )
    AGG_METHODS = (
        ('SUM', 'Sum()'),
        # ('COUNT', 'Count()'), # eg. we have a facility reporting patient aggregates but also a facility level field (has_incinerator)
        # ('AVG', 'Average()'), # needs a corresponding 'population' data element that provides the ratios when we combine two (or more) averages
    )
    
    dhis2_uid = models.CharField(max_length=11, blank=True, null=True) #TODO: add unique check and check for a minimum length of 11 as well
    name = models.CharField(max_length=128, unique=True)
    alias = models.CharField(max_length=128, blank=True, null=True)
    value_type = models.CharField(max_length=8, choices=VALUE_TYPES)
    value_min = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Minimum Value', blank=True, null=True)
    value_max = models.DecimalField(max_digits=17, decimal_places=4, verbose_name='Maximum Value', blank=True, null=True)
    aggregation_method = models.CharField(max_length=8, choices=AGG_METHODS)

    def validate_unique(self, exclude=None):
        super(DataElement, self).validate_unique(exclude=exclude)

        # name already exists as an alias
        if DataElement.objects.filter(Q(alias__iexact=self.name), ~Q(id=self.id)).exists():
            raise ValidationError({'name': 'Name already used as an alias: \'%s\'' % (self.name,)})
        if self.alias:
            # alias already exists as a name/alias
            if DataElement.objects.filter(Q(name__iexact=self.alias)|Q(alias__iexact=self.alias), ~Q(id=self.id)).exists():
                raise ValidationError({'alias': 'Alias already used as a name/alias: \'%s\'' % (self.alias,)})

    def save(self, *args, **kwargs):
        self.validate_unique()
        super(DataElement, self).save(*args, **kwargs)

    def __repr__(self):
        return 'DataElement<%s>' % (str(self),)

    def __str__(self):
        return '%s' % (self.name,)

class Category(models.Model):
    name = models.CharField(max_length=128, unique=True)

    class Meta:
        verbose_name_plural = 'categories'

    def __str__(self):
        return self.name

class CategoryCombo(models.Model):
    name = models.CharField(max_length=512)
    categories = models.ManyToManyField(Category)

    @classmethod
    def from_cat_names(cls, cat_names):
        sorted_names = sorted(cat_names) #TODO: sort based on the name of the classification the Category belongs to
        cat_list = [Category.objects.get_or_create(name=cat_name)[0] for cat_name in sorted_names]
        cc_name = '(%s)' % ', '.join(sorted_names)
        cat_combo, created = cls.objects.get_or_create(name=cc_name)
        if created:
            for categ in cat_list:
                cat_combo.categories.add(categ)
            cat_combo.save()

        return cat_combo

    def __str__(self):
        return self.name


# TODO: Consider tracking which data element each subcategory is from (reduce false matches and other? benefits)
CATEGORIES = [
    'Male',
    'Female',
    '18 Mths-<5 Years',
    '5-<10 Years',
    '10-<15 Years',
    '15-<19 Years',
    '19-<49 Years',
    '>49 Years',
    '10-19 Years',
    '20-24 Years',
    '>=25 Years',

    '<2 Years',
    '2 - < 5 Years (HIV Care)',
    '5 - 14 Years',
    '< 15 Years',
    '15 Years and above',
    # HMIS 106a: 1B ART QUARTERLY COHORT ANALYSIS REPORT: FOLLOW-UP
    'Alive on ART in Cohort',
    'Died',
    'Lost  to Followup',
    'Lost',
    'Started on ART-Cohort',
    'Stopped',
    'Transfered In',
    'Transferred Out',
    # HMIS 105: 1.3 OPD
    '0-28 Days',
    '29 Days-4 Years',
    '5-59 Years',
    '60andAbove Years',

    # VMMC
    '2<5 Years',
    '5-<15 Years',
    '15-49 Years',

    # HMIS 105 Laboratory
    'Under 5 years',
    '5 years and above',

    '<15',
    '15+',
]

import re
SEP_REGEX = '[\s,]+' # one or more of these characters in sequence
CATEGORY_REGEX = '|'.join('%s?(%s)' % (SEP_REGEX, re.escape(categ)) for categ in CATEGORIES)
SEXLESS_CATEGORY_REGEX = '|'.join('%s?(%s)' % (SEP_REGEX, re.escape(categ)) for categ in CATEGORIES[2:]) #TODO: even more horrible a hack

ICKY_CATEGS = (
    'Number of Male',
    'Male partners',
)

def unpack_data_element(de_long):
    if any([de_long.upper().find(s.upper()) >=0 for s in ICKY_CATEGS]):
        m = re.split(SEXLESS_CATEGORY_REGEX, de_long)
    else:
        m = re.split(CATEGORY_REGEX, de_long)
    # squash list of matches by removing blank and None entries (and False and numeric zeroes)
    de_name, *category_list = tuple(filter(None, m))
    cat_str = ', '.join(category_list)

    # deals with cases where the data element name includes a subcategory ('105-2.1a Male partners received HIV test results in eMTCT')
    # and matches multiple subcategories ('Lost' and 'Lost  to Followup' in '106a Cohort  All patients 12 months Lost  to Followup')
    #TODO: reimplement this, it is a really ugly hack
    if any([cat not in CATEGORIES for cat in category_list]):
        cat_str = ' '.join(category_list)
        if cat_str not in CATEGORIES:
            de_name = de_long
            cat_str = ''

    de_instance, created = DataElement.objects.get_or_create(name=de_name, value_type='NUMBER', value_min=None, value_max=None, aggregation_method='SUM')
    if len(category_list):
        return (de_instance, CategoryCombo.from_cat_names(category_list))
    else:
        return (de_instance, None)

class DataValueQuerySet(models.QuerySet):
    """Convenience queryset methods for handling datavalues"""
    def what(self, *names):
        de_filters = None
        if names:
            for de in names:
                if de is None:
                    continue # skip any names/uids with value of None
                if de_filters:
                    de_filters = de_filters | Q(data_element__name__iexact=de) | Q(data_element__alias__iexact=de)
                else:
                    de_filters = Q(data_element__name__iexact=de) | Q(data_element__alias__iexact=de)

        qs = self.annotate(de_name=F('data_element__name'))
        qs = qs.annotate(de_uid=F('data_element__dhis2_uid'))
        if de_filters:
            qs = qs.filter(de_filters)
        return qs

    def where(self):
        raise NotImplementedError()

    def when(self):
        raise NotImplementedError()

class DataValueManager(models.Manager):
    """Attach our custom queryset methods to the model manager"""
    def get_queryset(self):
        return DataValueQuerySet(self.model, using=self._db)

    def what(self, *names):
        return self.get_queryset().what(*names)

    def where(self):
        raise NotImplementedError()

    def when(self):
        raise NotImplementedError()

def get_default_category_combo():
    return CategoryCombo.objects.get(id=1)

class DataValue(models.Model):
    data_element = models.ForeignKey(DataElement, related_name='data_values')
    category_combo = models.ForeignKey(CategoryCombo, related_name='data_values', default=1)
    site_str = models.CharField(max_length=128)
    org_unit = models.ForeignKey(OrgUnit, related_name='data_values')
    numeric_value = models.DecimalField(max_digits=17, decimal_places=4)
    month = models.CharField(max_length=7, blank=True, null=True) # ISO 8601 format '2017-09'
    quarter = models.CharField(max_length=7, blank=True, null=True) # ISO 8601 format '2017-Q3'
    year = models.CharField(max_length=4, blank=True, null=True) # ISO 8601 format '2017'
    source_doc = models.ForeignKey(SourceDocument, related_name='data_values')

    objects = DataValueManager() # override the default manager

    class Meta():
        unique_together = (('data_element', 'category_combo', 'org_unit', 'year', 'quarter', 'month'),)

    def __repr__(self):
        return 'DataValue<%s [%s], %s, %s, %d>' % (str(self.data_element), self.category_combo, self.site_str,  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

    def __str__(self):
        return '%s [%s], %s, %s, %d' % (str(self.data_element), self.category_combo, self.site_str.split(' => ')[-1],  next(filter(None, (self.month, self.quarter, self.year))), self.numeric_value,)

@lru_cache(maxsize=16) # memoize to reduce cost of "parsing"
def extract_periods(period_str):
    from .grabbag import period_to_dates, dates_to_iso_periods
    dates = period_to_dates(period_str)
    return dates_to_iso_periods(*dates)

def load_excel_to_datavalues(source_doc, max_sheets=4):
    from collections import defaultdict
    import re
    import calendar
    import openpyxl

    MONTH_REGEX = r'[\s]*(%s) [0-9]{4}[\s]*' % ('|'.join(calendar.month_name[1:]),)
    MONTH_PREFIX_REGEX = r'^[\s]*(%s) ([0-9]{4})?[\s]*' % ('|'.join(calendar.month_name[1:]),)

    DE_COLUMN_START = 4 # 0-based index of first dataelement column in worksheet

    wb = openpyxl.load_workbook(source_doc.file.path)
    logger.debug(wb.get_sheet_names())

    wb_loc_values = defaultdict(list) # when a new key is encountered return a new empty list

    for ws_name in wb.get_sheet_names()[:max_sheets]: #['Step1', 'Targets']:
        if ws_name in ['Validations']:
            continue
        ws = wb[ws_name]
        logger.debug((ws_name, ws.max_row, ws.max_column))

        headers = [cell.value for cell in ws.rows[0]]
        # discard the month (and space) prefix on the data element names
        clean_headers = (re.sub(MONTH_PREFIX_REGEX, '', h) for h in headers[DE_COLUMN_START:] if h is not None)
        data_elements = tuple(unpack_data_element(de) for de in clean_headers)


        for row in ws.rows[1:]: # skip header row
            period, *location_parts = [c.value for c in row[:DE_COLUMN_START]]
            if not period or not any(location_parts):
                continue # ignore rows where period or location is missing
            iso_year, iso_quarter, iso_month = extract_periods(str(period).strip())
            location_parts = ('Uganda', *filter(None, location_parts)) # turn to tuple and prepend name of root OrgUnit
            current_ou = OrgUnit.from_path_recurse(*location_parts)
            location = ' => '.join(location_parts)
            logger.debug((period, location))

            site_val_cells = row[DE_COLUMN_START:]
            site_values = zip(data_elements, (c.value for c in site_val_cells))
            dv_construct = partial(DataValue, site_str=location, org_unit=current_ou, month=iso_month, quarter=iso_quarter, year=iso_year, source_doc=source_doc)
            data_values = list()
            for (de, cc), dv in site_values:
                if dv is None or (isinstance(dv, str) and dv.strip() == ''):
                    continue # skip rows with empty values
                if cc:
                    data_values.append(dv_construct(data_element=de, category_combo=cc, numeric_value=Decimal(dv)))
                else:
                    data_values.append(dv_construct(data_element=de, numeric_value=Decimal(dv)))

            wb_loc_values[location].extend(data_values)

    return dict(wb_loc_values) # convert back to a normal dict for our callers

def de_pivot_col(de):
    return 'DE_%d' % (de.id,)

def pivot_clause(data_elements):
    return ',\n'.join("SUM(CASE WHEN de.name = '%s' THEN dv.numeric_value ELSE 0 END) as '%s'" % (de.name, de_pivot_col(de)) for de in data_elements)

def validation_expr(left, right, operator):
    pass

def validation_expr_elements(expr):
    names = list()
    for de_tup in DataElement.objects.all().values_list('name', 'alias'):
        names.extend(de_tup)
    names = filter(None, names)
    sorted_names = list(sorted(names, reverse=True)) # sort puts longest matches first
    DE_REGEX = '|'.join('%s' % (re.escape(de_name),) for de_name in sorted_names)
    m = re.findall(DE_REGEX, expr, flags=re.IGNORECASE)
    print(m)
    return tuple(filter(None, m))

def load_excel_to_validations(source_doc):
    import openpyxl

    wb = openpyxl.load_workbook(source_doc.file.path) #TODO: ensure we close the workbook file. use a context manager?
    logger.debug(wb.get_sheet_names())

    for ws_name in wb.get_sheet_names():
        if ws_name != 'Validations':
            continue

        ws = wb[ws_name]
        logger.debug((ws_name, ws.max_row, ws.max_column))

        
        for row in ws.rows[1:]: # skip header row
            validation_name, l_exp, op, r_exp, *_ = [c.value for c in row]
            if not l_exp or not op or not r_exp:
                continue # ignore rows where any part of the rule is missing
            print(validation_name, l_exp, op, r_exp)
            l_element_names = validation_expr_elements(l_exp)
            r_element_names = validation_expr_elements(r_exp)
            element_names = l_element_names + r_element_names
            bad_rules = ['Mal_1', 'Mal_6', 'Mal_7', 'Mal_11']
            if len(element_names) > 0 and len(l_element_names) > 0 and len(r_element_names) > 0 and validation_name not in bad_rules: #TODO: exclude dodgy rule for demo
                try:
                    vr = ValidationRule.objects.get(name=validation_name)
                    vr.left_expr, vr.right_expr, vr.operator = l_exp, r_exp, op
                except ValidationRule.DoesNotExist as e:
                    vr = ValidationRule(name=validation_name, left_expr=l_exp, right_expr=r_exp, operator=op)
                vr.save()
                print(vr.view_name())

    return

    ou_level = month_multiple = None
    search_periods = ['2016']
    de_meta_list = query_de_meta(tt_names)
    print(de_meta_list)
    calc_exprs = (
        ('DE_5*100/DE_22', ['DE_22',]),
        ('DE_6*100/DE_22', ['DE_22',]),
    )
    if ou_level is None: # if not explicitly given use the highest orgunit level
        ou_level = min(map(lambda x: x.ou_level, de_meta_list))
    if month_multiple is None:
        month_multiple = max(map(lambda x: x.month_multiple, de_meta_list))
    calc_query = mk_calculation_sql(calc_exprs, de_meta_list, [], ou_level, search_periods, month_multiple)
    print(calc_query)

def mk_validation_rule_sql(rule_expr, data_elements):
    de_meta_list = query_de_meta(data_elements)
    ou_level = min(map(lambda x: x.ou_level, de_meta_list))
    month_multiple = max(map(lambda x: x.month_multiple, de_meta_list))

    subst_rule_expr = rule_expr
    for de_meta in sorted(de_meta_list, key=lambda x: x.name, reverse=True):
        subst_rule_expr = re.sub('(?i)' + re.escape(de_meta.name), 'DE_%d' % (de_meta.id,), subst_rule_expr)
        if de_meta.alias:
            subst_rule_expr = re.sub('(?i)' + re.escape(de_meta.alias), 'DE_%d' % (de_meta.id,), subst_rule_expr)

    return mk_calculation_sql([(subst_rule_expr, [])], de_meta_list, [], ou_level, [], month_multiple)

def query_de_meta(de_names):
    """
    Given a sequence of dataelement names, return a corresponding sequence of
    tuples containing the name, id, highest orgunit level it is collected at,
    and the largest period type it is collected for (as a multiple of month)
    """
    from functools import reduce
    from collections import namedtuple

    if len(de_names) == 0:
        return tuple()
    
    q_objs = reduce(lambda x, y: x | y, (Q(alias__iexact=de_name)|Q(name__iexact=de_name) for de_name in de_names))
    qs = DataElement.objects.filter(q_objs)
    qs = qs.annotate(ou_level=Min(F('data_values__org_unit__level')))
    qs = qs.annotate(month_multiple=Min(Case(When(data_values__month__isnull=False, then=1), When(data_values__quarter__isnull=False, then=4), When(data_values__year__isnull=False, then=12), default=None, output_field=models.IntegerField())))
    qs = qs.order_by('name', 'id', 'ou_level', 'month_multiple')
    
    DataElementMeta = namedtuple('DataElementMeta', ['name', 'alias', 'id', 'ou_level', 'month_multiple'])
    return tuple(DataElementMeta(**v) for v in qs.values('name', 'alias', 'id', 'ou_level', 'month_multiple'))

def fields_for_ou_level(ou_level):
    return ('country', 'district', 'subcounty', 'facility')[:ou_level+1]

def fields_for_month_multiple(month_mul):
    return ('year', 'quarter', 'month')[:(12, 3, 1).index(month_mul)+1]

def gen_pairs(iterable):
    item2_iter = iter(iterable)
    _ = next(item2_iter) # skip one ahead
    for item1 in iterable:
        yield (item1, next(item2_iter))

def mk_de_group_sql(de_meta_list, all_fields, ou_level):
    select_clause = ' '.join(['SELECT', ', '.join(all_fields)])
    
    _tables = ['cannula_datavalue dv', 'cannula_orgunit ou', 'cannula_dataelement de'] + ['cannula_orgunit ou%d' % i for i in range(ou_level+1)]
    from_clause = ' '.join(['FROM', ', '.join(_tables)])

    join_filter_subclause = ' AND '.join(('dv.org_unit_id=ou.id', 'dv.data_element_id=de.id'))
    ou_traversals = ['ou0.parent_id IS NULL'] + ['ou%d.parent_id=ou%d.id' % (c, p) for p,c in gen_pairs(range(ou_level+1))]
    ou_traversals.append('ou.id = ou%d.id' % (ou_level,))
    de_filters = ['de.name=\'%s\'' % (de_m.name,) for de_m in de_meta_list] #TODO: switch to data element IDs/UIDs/CODEs to avoid SQL injection
    de_filter_clause = '(%s)' % ('\nOR '.join(de_filters),)
    where_parts = [join_filter_subclause, *ou_traversals, de_filter_clause]
    where_clause = 'WHERE ' + ('\nAND '.join(where_parts))

    return '\n'.join([select_clause, from_clause, where_clause])

def mk_union_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    from itertools import groupby

    ou_fields = fields_for_ou_level(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    print(ou_fields, period_fields)

    hier_ou_pairs = tuple(('ou%d' % (i,), f) for i, f in enumerate(ou_fields))
    hier_ou_fields = tuple('%s.name as %s' % (code, desc) for code, desc in hier_ou_pairs)

    placeholder_fields = ', '.join(['NULL as %s' % (f,) for f in (period_fields + ou_fields)])
    union_parts = ['SELECT %s, NULL as de_name, NULL as numeric_value FROM cannula_datavalue dv' % (placeholder_fields,)]

    grouped_de_metas = groupby(de_meta_list, lambda x: (x.ou_level, x.month_multiple))
    for g in grouped_de_metas:
        g_ident, g_seq = g
        g_seq = list(g_seq)
        g_ou_level, g_month_multiple = g_ident

        if g_month_multiple > period_month_multiple:
            my_period_fields = fields_for_month_multiple(g_month_multiple)
            my_periods = [tuple(filter(None, grabbag.dates_to_iso_periods(*grabbag.period_to_dates(p)))) for p in period_list]
            my_periods = [tuple('\'%s\'' % (p,) for p in p_tup) for p_tup in my_periods] #TODO: do proper db quoting
            print(my_periods)
            for p_tup in my_periods:
                all_fields = my_period_fields + p_tup[len(my_period_fields):] + hier_ou_fields + ('de.name as de_name' , 'numeric_value')
                group_select = mk_de_group_sql(g_seq, all_fields, g_ou_level)
                union_parts.append(group_select)
        else:
            my_period_fields = period_fields
            all_fields = my_period_fields + hier_ou_fields + ('de.name as de_name' , 'numeric_value')
            group_select = mk_de_group_sql(g_seq, all_fields, g_ou_level)
            union_parts.append(group_select)

    return '\nUNION ALL\n'.join(union_parts)

def mk_aggregate_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    union_sql = mk_union_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = fields_for_ou_level(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    groupby_fields = period_fields+ou_fields+('de_name',)
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+('sum(numeric_value) as numeric_sum', 'count(numeric_value) as numeric_count'))])
    group_order_clause = 'AS q_aggregate\nGROUP BY %s\nORDER BY %s' % (groupby_fields_str, groupby_fields_str)

    aggregate_sql = select_clause + '\n' + 'FROM (' + '\n' + union_sql + '\n' + ') ' + group_order_clause

    return aggregate_sql

def mk_pivot_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    aggregate_sql = mk_aggregate_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = fields_for_ou_level(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    pivot_fields = list()
    for de in de_meta_list:
        if de.month_multiple <= period_month_multiple:
            de_pivot_str = 'SUM(CASE WHEN de_name = \'%s\' THEN numeric_sum ELSE 0 END) as DE_%d' % (de.name, de.id)
        else:
            de_pivot_str = 'SUM(CASE WHEN de_name = \'%s\' THEN numeric_sum/%f ELSE 0 END) as DE_%d' % (de.name, de.month_multiple/period_month_multiple, de.id)
        pivot_fields.append(de_pivot_str)
    groupby_fields = period_fields + ou_fields
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+tuple(pivot_fields))])
    group_clause = 'AS q_pivot\nGROUP BY %s' % (groupby_fields_str)

    pivot_sql = select_clause + '\n' + 'FROM (' + '\n' + aggregate_sql + '\n' + ') ' + group_clause

    return pivot_sql

def mk_calc_fields(calculations):
    calc_fields = list()
    for i, (calc_exp, zero_checks) in enumerate(calculations, start=1):
        z_c_str = ' AND '.join('(%s != 0)' % (z_c_field) for z_c_field in zero_checks)
        if len(zero_checks) > 0:
            calc_str = 'CASE WHEN %s THEN %s ELSE NULL END as DE_CALC_%d' % (z_c_str, calc_exp, i)
        else:
            calc_str = '%s as DE_CALC_%d' % (calc_exp, i)
        calc_fields.append(calc_str)

    return tuple(calc_fields)

def mk_calculation_sql(calculations, de_meta_list, ou_list, ou_level, period_list, period_month_multiple):
    from collections import defaultdict

    print('OU_PARAM: %d, PERIOD_PARAM: %d' % (ou_level, period_month_multiple))
    
    pivot_sql = mk_pivot_sql(de_meta_list, ou_list, ou_level, period_list, period_month_multiple)
    ou_fields = fields_for_ou_level(ou_level)
    period_fields = fields_for_month_multiple(period_month_multiple)
    calc_src_fields =  tuple('DE_%d' % (de.id,) for de in de_meta_list)
    calc_fields = mk_calc_fields(calculations)
    groupby_fields = period_fields + ou_fields
    groupby_fields_str = ', '.join(groupby_fields)
    select_clause = ' '.join(['SELECT', ', '.join(groupby_fields+calc_src_fields+calc_fields)])

    where_groups = defaultdict(list)
    p_fields = fields_for_month_multiple(period_month_multiple)
    for p in period_list:
        p_vals = tuple(filter(None, grabbag.dates_to_iso_periods(*grabbag.period_to_dates(p))))
        for p_pair in zip(p_fields, p_vals):
            if p_pair not in where_groups[p_pair[0]]:
                where_groups[p_pair[0]].append(p_pair)
    where_parts = ['(%s)' % (' OR '.join('%s=\'%s\'' % (f, v) for f, v in l),) for k, l in where_groups.items()]

    if len(where_parts) > 0:
        where_clause = 'WHERE (%s)' % ' AND '.join(where_parts)
    else:
        where_clause = ''

    calculation_sql = select_clause + '\n' + 'FROM (' + '\n' + pivot_sql + '\n' + ') AS q_calculate' + '\n' + where_clause

    return calculation_sql

class ValidationRule(models.Model):
    name = models.CharField(max_length=128, unique=True)
    left_expr = models.CharField(max_length=256) #TODO: store the cleaned up expression with symbolic data element names
    right_expr = models.CharField(max_length=256)
    operator = models.CharField(max_length=2) #TODO: make this a choice field
    #TODO: add a description/comments field ?

    data_elements = models.ManyToManyField(DataElement)

    def expression(self):
        return ' '.join([self.left_expr, self.operator, self.right_expr])

    def view_name(self):
        return 'vw_validation_%d' % (self.id,)

    def save(self, *args, **kwargs):
        from django.db import connection

        super(ValidationRule, self).save(*args, **kwargs)
        
        # parse and collect data element names
        l_element_names = validation_expr_elements(self.left_expr)
        r_element_names = validation_expr_elements(self.right_expr)
        element_names = l_element_names + r_element_names
        
        # modify list of data elements
        new_meta_list = query_de_meta(element_names)
        curr_meta_list = query_de_meta(self.data_elements.all().values_list('name'))
        for de_meta in curr_meta_list: # remove, if not in new list
            if de_meta not in new_meta_list:
                self.data_elements.remove(DataElement.objects.get(id=de_meta.id))
        for de_meta in new_meta_list: # add, if not in current list
            if de_meta not in curr_meta_list:
                self.data_elements.add(DataElement.objects.get(id=de_meta.id))

        super(ValidationRule, self).save(*args, **kwargs)

        # create the view
        sql = mk_validation_rule_sql(self.expression(), element_names)
        view_sql = 'CREATE OR REPLACE VIEW %s AS\n%s' % (self.view_name(), sql)
        cursor = connection.cursor()
        cursor.execute(view_sql, [])

    def __str__(self):
        return self.name
        